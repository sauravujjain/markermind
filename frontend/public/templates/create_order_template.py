#!/usr/bin/env python3
"""Create order template Excel file matching the expected format."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_order_template():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers matching expected format
    # Fixed columns + dynamic size columns
    FIXED_HEADERS = ["Order No.", "Style No.", "Fabric", "Order Color", "Extra %"]

    # Example data from user's order_data.xlsx
    # Order 1: 4 colors, 3 fabrics, 7 sizes (46-58)
    # Order 2: 1 color, 1 fabric, 3 sizes (S, M, L)

    ORDER1_SIZES = ["46", "48", "50", "52", "54", "56", "58"]
    ORDER2_SIZES = ["S", "M", "L"]
    ALL_SIZES = ORDER1_SIZES + ORDER2_SIZES  # Union for template

    EXAMPLE_DATA = [
        # Order 1: style1, 3 fabrics (SO1, SO2, FO1), 4 colors each
        ["Order1", "style1", "SO1", "8320", 0, 74, 244, 347, 342, 265, 162, 62, "", "", ""],
        ["Order1", "style1", "SO1", "8535", 0, 23, 112, 172, 166, 114, 74, 17, "", "", ""],
        ["Order1", "style1", "SO1", "8820", 0, 29, 172, 248, 254, 191, 145, 45, "", "", ""],
        ["Order1", "style1", "SO1", "9990", 0, 20, 104, 167, 162, 114, 78, 33, "", "", ""],
        ["Order1", "style1", "SO2", "8320", 0, 74, 244, 347, 342, 265, 162, 62, "", "", ""],
        ["Order1", "style1", "SO2", "8535", 0, 23, 112, 172, 166, 114, 74, 17, "", "", ""],
        ["Order1", "style1", "SO2", "8820", 0, 29, 172, 248, 254, 191, 145, 45, "", "", ""],
        ["Order1", "style1", "SO2", "9990", 0, 20, 104, 167, 162, 114, 78, 33, "", "", ""],
        ["Order1", "style1", "FO1", "8320", 0, 74, 244, 347, 342, 265, 162, 62, "", "", ""],
        ["Order1", "style1", "FO1", "8535", 0, 23, 112, 172, 166, 114, 74, 17, "", "", ""],
        ["Order1", "style1", "FO1", "8820", 0, 29, 172, 248, 254, 191, 145, 45, "", "", ""],
        ["Order1", "style1", "FO1", "9990", 0, 20, 104, 167, 162, 114, 78, 33, "", "", ""],
        # Order 2: style2, 1 fabric (Shell), 1 color (Red), 3% extra
        ["Order2", "style2", "Shell", "Red", 3, "", "", "", "", "", "", "", 200, 250, 225],
    ]

    # ============ Sheet 1: Template ============
    ws_template = wb.active
    ws_template.title = "Template"

    # Headers
    headers = FIXED_HEADERS + ALL_SIZES
    for col, header in enumerate(headers, 1):
        cell = ws_template.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Empty data rows (2-20)
    for row in range(2, 21):
        for col in range(1, len(headers) + 1):
            cell = ws_template.cell(row=row, column=col, value="")
            cell.border = thin_border
            if col <= 4:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws_template.column_dimensions['A'].width = 12  # Order No.
    ws_template.column_dimensions['B'].width = 12  # Style No.
    ws_template.column_dimensions['C'].width = 10  # Fabric
    ws_template.column_dimensions['D'].width = 14  # Order Color
    ws_template.column_dimensions['E'].width = 10  # Extra %
    for col in range(6, len(headers) + 1):
        ws_template.column_dimensions[get_column_letter(col)].width = 6

    # ============ Sheet 2: Example Data ============
    ws_example = wb.create_sheet("Example Data")

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws_example.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, row_data in enumerate(EXAMPLE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_example.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
            cell.border = thin_border
            if col_idx <= 4:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws_example.column_dimensions['A'].width = 12
    ws_example.column_dimensions['B'].width = 12
    ws_example.column_dimensions['C'].width = 10
    ws_example.column_dimensions['D'].width = 14
    ws_example.column_dimensions['E'].width = 10
    for col in range(6, len(headers) + 1):
        ws_example.column_dimensions[get_column_letter(col)].width = 6

    # Add summary section
    summary_row = len(EXAMPLE_DATA) + 3
    ws_example.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=12)

    ws_example.cell(row=summary_row + 1, column=1, value="Order 1 (style1):").font = Font(bold=True)
    ws_example.cell(row=summary_row + 2, column=1, value="  • 3 Fabrics: SO1, SO2, FO1 (Self Outer, Self Outer, Fusing Outer)")
    ws_example.cell(row=summary_row + 3, column=1, value="  • 4 Colors: 8320, 8535, 8820, 9990")
    ws_example.cell(row=summary_row + 4, column=1, value="  • 7 Sizes: 46, 48, 50, 52, 54, 56, 58")
    ws_example.cell(row=summary_row + 5, column=1, value="  • 12 lines total (3 fabrics × 4 colors)")

    ws_example.cell(row=summary_row + 7, column=1, value="Order 2 (style2):").font = Font(bold=True)
    ws_example.cell(row=summary_row + 8, column=1, value="  • 1 Fabric: Shell")
    ws_example.cell(row=summary_row + 9, column=1, value="  • 1 Color: Red")
    ws_example.cell(row=summary_row + 10, column=1, value="  • 3 Sizes: S, M, L")
    ws_example.cell(row=summary_row + 11, column=1, value="  • 3% Extra buffer")

    ws_example.cell(row=summary_row + 13, column=1, value="Notes:").font = Font(bold=True)
    ws_example.cell(row=summary_row + 14, column=1, value="  • Each row is a unique (Order + Fabric + Color) combination")
    ws_example.cell(row=summary_row + 15, column=1, value="  • Size columns can vary - only include sizes relevant to your order")
    ws_example.cell(row=summary_row + 16, column=1, value="  • Empty size cells are treated as 0 quantity")
    ws_example.cell(row=summary_row + 17, column=1, value="  • Extra % adds buffer to quantities for waste/defects")

    # Save
    output_path = "/home/sarv/projects/MarkerMind/frontend/public/templates/order_template.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

if __name__ == "__main__":
    create_order_template()
