from typing import List
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session, joinedload
import io
import csv

from ...database import get_db
from ...models import User, Cutplan, CutplanMarker, Order, Pattern, MarkerLayout, PatternFabricMapping, TestMarkerResult
from ..deps import get_current_user

router = APIRouter(prefix="/export", tags=["exports"])


@router.get("/cutplan/{cutplan_id}/docket")
async def export_cutting_docket(
    cutplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export cutting docket as PDF."""
    cutplan = db.query(Cutplan).join(Order).filter(
        Cutplan.id == cutplan_id,
        Order.customer_id == current_user.customer_id
    ).first()
    if not cutplan:
        raise HTTPException(status_code=404, detail="Cutplan not found")

    # TODO: Generate PDF using ReportLab or WeasyPrint
    # For now, return a simple text representation
    content = f"""
CUTTING DOCKET
==============
Order: {cutplan.order.order_number}
Cutplan: {cutplan.name}
Status: {cutplan.status}

SUMMARY
-------
Total Yards: {cutplan.total_yards:.2f}
Unique Markers: {cutplan.unique_markers}
Total Plies: {cutplan.total_plies}
Total Cost: ${cutplan.total_cost:.2f}

MARKERS
-------
"""
    for marker in cutplan.markers:
        content += f"""
Ratio: {marker.ratio_str}
Efficiency: {marker.efficiency:.1f}%
Length: {marker.length_yards:.2f} yards
Total Plies: {marker.total_plies}
Cuts: {marker.cuts}
Plies by Color: {marker.plies_by_color}
---
"""

    return Response(
        content=content,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename=docket_{cutplan_id}.txt"}
    )


@router.get("/cutplan/{cutplan_id}/csv")
async def export_cutplan_csv(
    cutplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export cutplan as CSV."""
    cutplan = db.query(Cutplan).join(Order).filter(
        Cutplan.id == cutplan_id,
        Order.customer_id == current_user.customer_id
    ).first()
    if not cutplan:
        raise HTTPException(status_code=404, detail="Cutplan not found")

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Marker #",
        "Ratio",
        "Efficiency %",
        "Length (yards)",
        "Total Plies",
        "Cuts",
        "Plies by Color"
    ])

    # Data
    for i, marker in enumerate(cutplan.markers, 1):
        plies_str = ", ".join(f"{k}:{v}" for k, v in (marker.plies_by_color or {}).items())
        writer.writerow([
            i,
            marker.ratio_str,
            f"{marker.efficiency:.1f}" if marker.efficiency else "",
            f"{marker.length_yards:.2f}" if marker.length_yards else "",
            marker.total_plies or "",
            marker.cuts or "",
            plies_str
        ])

    # Summary row
    writer.writerow([])
    writer.writerow(["SUMMARY"])
    writer.writerow(["Total Yards", f"{cutplan.total_yards:.2f}" if cutplan.total_yards else ""])
    writer.writerow(["Total Plies", cutplan.total_plies or ""])
    writer.writerow(["Unique Markers", cutplan.unique_markers or ""])
    writer.writerow(["Total Cost", f"${cutplan.total_cost:.2f}" if cutplan.total_cost else ""])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=cutplan_{cutplan_id}.csv"}
    )


@router.get("/markers/{marker_id}/dxf")
async def export_marker_dxf(
    marker_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export marker layout as DXF."""
    raise HTTPException(status_code=501, detail="DXF export not yet implemented")


@router.get("/test-marker/{result_id}/dxf")
async def export_test_marker_dxf(
    result_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export a saved test marker result as DXF (generated at nest time via export_marker_dxf)."""
    result = db.query(TestMarkerResult).filter(
        TestMarkerResult.id == result_id,
        TestMarkerResult.created_by == current_user.id
    ).first()
    if not result:
        raise HTTPException(status_code=404, detail="Test marker result not found")
    if not result.dxf_data:
        raise HTTPException(status_code=400, detail="No DXF data stored for this result")

    filename = f"marker_{result.ratio_str}_{result.efficiency*100:.0f}pct.dxf"
    return Response(
        content=result.dxf_data,
        media_type="application/dxf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/order/{order_id}/summary")
async def export_order_summary(
    order_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export order summary including all cutplans."""
    order = db.query(Order).filter(
        Order.id == order_id,
        Order.customer_id == current_user.customer_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    output = io.StringIO()
    writer = csv.writer(output)

    # Order info
    writer.writerow(["ORDER SUMMARY"])
    writer.writerow(["Order Number", order.order_number])
    writer.writerow(["Status", order.status])
    writer.writerow([])

    # Demand
    writer.writerow(["DEMAND"])
    size_codes = set()
    for color in order.colors:
        for sq in color.size_quantities:
            size_codes.add(sq.size_code)
    size_codes = sorted(size_codes)

    writer.writerow(["Color"] + size_codes + ["Total"])
    for color in order.colors:
        qty_by_size = {sq.size_code: sq.quantity for sq in color.size_quantities}
        row = [color.color_code]
        total = 0
        for size in size_codes:
            qty = qty_by_size.get(size, 0)
            row.append(qty)
            total += qty
        row.append(total)
        writer.writerow(row)

    writer.writerow([])

    # Cutplans
    writer.writerow(["CUTPLANS"])
    writer.writerow(["Name", "Status", "Efficiency", "Yards", "Markers", "Cost"])
    for cp in order.cutplans:
        writer.writerow([
            cp.name or "",
            cp.status,
            f"{cp.efficiency:.1f}%" if cp.efficiency else "",
            f"{cp.total_yards:.2f}" if cp.total_yards else "",
            cp.unique_markers or "",
            f"${cp.total_cost:.2f}" if cp.total_cost else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=order_{order.order_number}_summary.csv"}
    )


@router.get("/order/{order_id}/excel")
async def export_order_excel(
    order_id: str,
    include_markers: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export all approved/refined cutplans as Excel workbook — one tab per fabric.
    Each fabric tab shows the demand table once, then each cutplan stacked with
    marker tables, totals, and fully recalculated cost breakdowns.

    Query params:
        include_markers: if true, embed marker PNG images after each marker row
    """
    import openpyxl

    from ...services.excel_export_service import (
        TITLE_FONT, SUBTITLE_FONT, CUTPLAN_FILL, HEADER_FONT,
        write_demand_table, write_marker_table,
        write_cost_breakdown, write_cost_charts,
        write_lay_plan_sheet, write_marker_details,
    )

    order = db.query(Order).filter(
        Order.id == order_id,
        Order.customer_id == current_user.customer_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Get canonical size ordering and perimeter data from pattern
    canonical_sizes = None
    perimeter_all_materials = {}
    pattern = None
    if order.pattern_id:
        pattern = db.query(Pattern).filter(Pattern.id == order.pattern_id).first()
        if pattern:
            if pattern.available_sizes:
                canonical_sizes = pattern.available_sizes
            meta = pattern.parse_metadata or {}
            perimeter_all_materials = meta.get("perimeter_by_size", {})

    # Load cost config for customer
    from ...models import CostConfig as CostConfigModel, Fabric as FabricModel
    cost_config = db.query(CostConfigModel).filter(
        CostConfigModel.customer_id == current_user.customer_id
    ).first()

    # Cost config defaults
    fabric_cost_per_yard = cost_config.fabric_cost_per_yard if cost_config else 3.0
    spreading_cost_per_yard = cost_config.spreading_cost_per_yard if cost_config else 0.00122
    spreading_cost_per_ply = cost_config.spreading_cost_per_ply if cost_config else 0.013
    cutting_labor_per_hour = cost_config.cutting_labor_cost_per_hour if cost_config else 1.0
    cutting_workers = cost_config.cutting_workers_per_cut if cost_config else 1
    cutting_speed_cm_s = cost_config.cutting_speed_cm_per_s if cost_config else 10.0
    cutting_cost_per_cm = (cutting_labor_per_hour * cutting_workers) / 3600.0 / cutting_speed_cm_s if cutting_speed_cm_s > 0 else 0.0
    prep_cost_per_meter = 0.0
    if cost_config:
        if cost_config.prep_perf_paper_enabled:
            prep_cost_per_meter += cost_config.prep_perf_paper_cost_per_m or 0.0
        if cost_config.prep_underlayer_enabled:
            prep_cost_per_meter += cost_config.prep_underlayer_cost_per_m or 0.0
        if cost_config.prep_top_layer_enabled:
            prep_cost_per_meter += cost_config.prep_top_layer_cost_per_m or 0.0
    else:
        prep_cost_per_meter = 0.25

    max_ply_height = cost_config.max_ply_height if cost_config else 100
    DEFAULT_PERIMETER_CM_PER_BUNDLE = 2540.0

    # Query ALL cutplans with status approved/refining/refined
    plans = db.query(Cutplan).filter(
        Cutplan.order_id == order_id,
        Cutplan.status.in_(["approved", "refined", "refining"]),
    ).options(
        joinedload(Cutplan.markers).joinedload(CutplanMarker.layout)
    ).all()

    if not plans:
        raise HTTPException(status_code=404, detail="No approved cutplans found for this order")

    # De-duplicate from joinedload
    seen_plan_ids = set()
    unique_plans = []
    for p in plans:
        if p.id not in seen_plan_ids:
            seen_plan_ids.add(p.id)
            unique_plans.append(p)
    plans = unique_plans

    # Group order lines by fabric code
    fabric_lines = {}
    for line in order.order_lines:
        fabric_lines.setdefault(line.fabric_code, []).append(line)

    # Collect all demand sizes across all fabrics
    demand_sizes_set = set()
    seen_colors = set()
    for line in order.order_lines:
        if line.color_code in seen_colors:
            continue
        seen_colors.add(line.color_code)
        for sq in line.size_quantities:
            if sq.quantity > 0:
                demand_sizes_set.add(sq.size_code)

    # Order sizes canonically
    if canonical_sizes:
        sizes = [s for s in canonical_sizes if s in demand_sizes_set]
        for s in sorted(demand_sizes_set):
            if s not in sizes:
                sizes.append(s)
    else:
        sizes = sorted(demand_sizes_set)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Unique fabric codes in order of appearance
    fabric_codes = list(dict.fromkeys(line.fabric_code for line in order.order_lines))

    # Look up fabric cost_per_yard from Fabric table
    fabric_cost_map = {}
    for fc in fabric_codes:
        fab = db.query(FabricModel).filter(
            FabricModel.customer_id == current_user.customer_id,
            FabricModel.code == fc,
        ).first()
        if fab and fab.cost_per_yard:
            fabric_cost_map[fc] = fab.cost_per_yard

    for fabric_code in fabric_codes:
        sheet_name = fabric_code[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Collect color demands for this fabric
        fabric_color_demands = {}
        fabric_seen_colors = set()
        for line in fabric_lines.get(fabric_code, []):
            if line.color_code in fabric_seen_colors:
                continue
            fabric_seen_colors.add(line.color_code)
            color_demand = {}
            for sq in line.size_quantities:
                if sq.quantity > 0:
                    color_demand[sq.size_code] = sq.quantity
            if color_demand:
                fabric_color_demands[line.color_code] = color_demand

        fc_cost_per_yard = fabric_cost_map.get(fabric_code, fabric_cost_per_yard)

        # Resolve perimeter_by_size for this fabric's material
        perimeter_by_size = {}
        if perimeter_all_materials and pattern:
            fab = db.query(FabricModel).filter(
                FabricModel.customer_id == current_user.customer_id,
                FabricModel.code == fabric_code,
            ).first()
            if fab:
                mapping = db.query(PatternFabricMapping).filter(
                    PatternFabricMapping.pattern_id == pattern.id,
                    PatternFabricMapping.fabric_id == fab.id,
                ).first()
                if mapping and mapping.material_name in perimeter_all_materials:
                    perimeter_by_size = perimeter_all_materials[mapping.material_name]
            if not perimeter_by_size and len(perimeter_all_materials) == 1:
                perimeter_by_size = list(perimeter_all_materials.values())[0]

        row = 1

        # -- Header --
        ws.cell(row=row, column=1, value=f"Order: {order.order_number}").font = TITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=f"Fabric: {fabric_code}")
        from openpyxl.styles import Font as XlFont
        cost_info = ws.cell(row=row, column=2, value=f"${fc_cost_per_yard:.2f}/yd")
        cost_info.font = XlFont(bold=True, size=11, color="006600")
        row += 2

        # -- Demand Table --
        row = write_demand_table(ws, row, sizes, fabric_color_demands)

        # -- Each cutplan stacked --
        for plan_idx, plan in enumerate(plans):
            plan_label = plan.name or f"Option {plan_idx + 1}"
            cell = ws.cell(row=row, column=1, value=f"CUTPLAN: {plan_label}")
            cell.font = SUBTITLE_FONT
            cell.fill = CUTPLAN_FILL
            for ci in range(2, 2 + len(sizes) + 7):
                ws.cell(row=row, column=ci).fill = CUTPLAN_FILL
            row += 1
            ws.cell(row=row, column=1, value=f"Status: {plan.status}")
            row += 2

            # -- Marker Table (always without images) --
            # Use max_ply_height from cutplan's solver_config if stored, else cost_config
            plan_max_ply = max_ply_height
            if plan.solver_config and isinstance(plan.solver_config, dict):
                plan_max_ply = plan.solver_config.get("max_ply_height", max_ply_height)

            cost_params = {
                "default_perimeter_cm": DEFAULT_PERIMETER_CM_PER_BUNDLE,
                "fc_cost_per_yard": fc_cost_per_yard,
                "spreading_cost_per_yard": spreading_cost_per_yard,
                "spreading_cost_per_ply": spreading_cost_per_ply,
                "cutting_cost_per_cm": cutting_cost_per_cm,
                "prep_cost_per_meter": prep_cost_per_meter,
            }
            row, cost_totals = write_marker_table(
                ws, row, sizes, plan.markers,
                perimeter_by_size, cost_params,
                max_ply_height=plan_max_ply,
            )

            # -- Summary --
            ws.cell(row=row, column=1, value=f"Total Yards: {cost_totals['total_fabric_yards']:.2f}").font = HEADER_FONT
            row += 2

            # -- Cost Breakdown + two pie charts --
            row, cost_data_start = write_cost_breakdown(ws, row, cost_totals, fc_cost_per_yard)
            write_cost_charts(ws, cost_data_start)
            row += 2

            # -- Marker Details with images (after full cutplan) --
            if include_markers:
                # Get fabric width from nesting jobs for this order
                from ...models import NestingJob as NestingJobModel
                nj = db.query(NestingJobModel).filter(
                    NestingJobModel.order_id == order_id,
                    NestingJobModel.status == "completed",
                ).first()
                fab_width = nj.fabric_width_inches if nj else None
                row = write_marker_details(ws, row, sizes, plan.markers, fab_width)
                row += 1

            # -- Lay Plan sheet for this fabric+cutplan --
            import re
            safe_fabric = re.sub(r'[\\/?*\[\]:]', '', fabric_code)[:10]
            safe_plan = re.sub(r'[\\/?*\[\]:]', '', plan_label)[:10]
            lay_name = f"Lay-{safe_fabric}-{safe_plan}"
            write_lay_plan_sheet(wb, lay_name, sizes, plan.markers, plan_max_ply)

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 20)

    # Write to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"order_{order.order_number}_cutplan.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
