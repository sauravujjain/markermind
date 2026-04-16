"""
Roll Plan API endpoints.

Provides Monte Carlo simulation and GA optimization for fabric roll usage
against an approved cutplan.
"""
from typing import Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload

import math
import statistics
import time
import traceback

from ...database import get_db, SessionLocal
from ...models import User, Cutplan, Order
from ...models.cutplan import CutplanStatus
from ...models.rollplan import (
    FabricRoll,
    RollPlan,
    RollPlanMode,
    RollPlanStatus,
    RollInputType,
)
from ...schemas.rollplan import (
    RollPlanCreateRequest,
    RollPlanResponse,
    RollPlanStatusResponse,
    RollPlanListItem,
    RollUploadResponse,
    FabricRollResponse,
    CutDocketResponse,
    TuneCutplanRequest,
    TuneStatusResponse,
    RollPreviewRow,
    RollPreviewResponse,
)
from ...services.rollplan_service import RollPlanService
from ..deps import get_current_user

router = APIRouter(prefix="/rollplans", tags=["rollplans"])
rollplan_service = RollPlanService()

# In-memory job tracking (matches cutplan pattern)
_rollplan_jobs: Dict[str, Dict] = {}
_tune_jobs: Dict[str, Dict] = {}  # rollplan_id -> tune job state


# ---------------------------------------------------------------------------
# Static / Preview endpoints (no auth required for template download)
# ---------------------------------------------------------------------------


@router.get("/sample-rolls-template")
async def download_sample_rolls_template():
    """Download a sample Excel template for roll inventory upload."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()

    # --- Sheet 1: Template ---
    ws = wb.active
    ws.title = "Template"
    mandatory_headers = ["Roll Number", "Roll Length"]
    optional_headers = ["Unit", "Roll Width", "Width Unit", "Shrinkage X%", "Shrinkage Y%", "Shade Group"]
    all_headers = mandatory_headers + optional_headers

    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        if header in mandatory_headers:
            cell.fill = yellow_fill

    # Placeholder row
    ws.cell(row=2, column=1, value="R001")
    ws.cell(row=2, column=2, value=100.0)
    ws.cell(row=2, column=3, value="yd")

    # Auto-width
    for col_idx in range(1, len(all_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    # --- Sheet 2: Sample Data ---
    ws2 = wb.create_sheet("Sample Data")
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        if header in mandatory_headers:
            cell.fill = yellow_fill

    sample_data = [
        ("R001", 105.2, "yd", 60, "in", None, None, "A"),
        ("R002", 98.7, "yd", 60, "in", None, None, "A"),
        ("R003", 112.0, "yd", None, None, None, None, None),
        ("R004", 87.5, "yd", 60, "in", 1.5, 0.8, "B"),
        ("R005", 95.3, "yd", None, None, None, None, "A"),
        ("R006", 120.0, "yd", 60, "in", None, None, None),
        ("R007", 80.1, "yd", None, None, 1.2, 0.5, "B"),
        ("R008", 110.4, "yd", 60, "in", None, None, "A"),
        ("R009", 92.8, "yd", None, None, None, None, None),
        ("R010", 103.6, "yd", 60, "in", 1.0, 0.6, "A"),
        ("R011", 88.9, "yd", None, None, None, None, "C"),
        ("R012", 115.7, "yd", 60, "in", None, None, None),
        ("R013", 97.2, "m", None, None, None, None, "B"),
        ("R014", 84.0, "yd", 60, "in", 1.3, 0.7, "A"),
        ("R015", 108.5, "yd", None, None, None, None, None),
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws2.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(all_headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=roll_inventory_template.xlsx"},
    )


@router.post("/parse-rolls-preview", response_model=RollPreviewResponse)
async def parse_rolls_preview(
    file: UploadFile = File(...),
    cutplan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Parse roll Excel and return summary + preview rows without saving to DB."""
    from ...services.rollplan_simulator import parse_roll_excel, PseudoRollConfig

    file_bytes = await file.read()
    try:
        rolls = parse_roll_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rolls:
        raise HTTPException(status_code=400, detail="No valid rolls found in file")

    lengths = [r.length_yards for r in rolls]
    median_len = round(statistics.median(lengths), 2)

    preview_rows = [
        RollPreviewRow(roll_number=r.roll_id, length_yards=r.length_yards)
        for r in rolls[:10]
    ]

    resp = RollPreviewResponse(
        rolls_count=len(rolls),
        total_length_yards=round(sum(lengths), 2),
        avg_length_yards=round(statistics.mean(lengths), 2),
        median_length_yards=median_len,
        min_length_yards=round(min(lengths), 2),
        max_length_yards=round(max(lengths), 2),
        preview_rows=preview_rows,
    )

    # Compute shortfall if cutplan_id given
    if cutplan_id:
        cutplan = db.query(Cutplan).join(Order).filter(
            Cutplan.id == cutplan_id,
            Order.customer_id == current_user.customer_id,
        ).first()
        if cutplan:
            total_fabric = 0.0
            if cutplan.markers:
                for cm in cutplan.markers:
                    ml = cm.length_yards or 0
                    pbc = cm.plies_by_color or {}
                    total_fabric += ml * sum(pbc.values())

            if total_fabric == 0:
                total_fabric = cutplan.total_yards or 0

            if total_fabric > 0:
                buffer_target = total_fabric * 1.05  # 5% buffer
                uploaded_total = sum(lengths)
                shortfall = max(0, buffer_target - uploaded_total)

                resp.fabric_required_yards = round(total_fabric, 2)
                if shortfall > 0:
                    resp.shortfall_yards = round(shortfall, 2)
                    synthetic_len = median_len
                    resp.synthetic_roll_length_yards = synthetic_len
                    resp.synthetic_rolls_needed = max(1, math.ceil(shortfall / synthetic_len)) if synthetic_len > 0 else 0
                else:
                    resp.shortfall_yards = 0

    return resp


# ---------------------------------------------------------------------------
# Background execution
# ---------------------------------------------------------------------------


def execute_rollplan_job(
    rollplan_id: str,
    ga_pop_size: int = 30,
    ga_generations: int = 50,
    confirm_shortfall: bool = False,
):
    """Execute roll plan simulation in the background."""
    db = SessionLocal()
    try:
        roll_plan = db.query(RollPlan).filter(RollPlan.id == rollplan_id).first()
        if not roll_plan:
            _rollplan_jobs[rollplan_id] = {
                "status": "failed", "progress": 0, "message": "Roll plan not found",
            }
            return

        roll_plan.status = RollPlanStatus.running
        roll_plan.progress = 0
        roll_plan.progress_message = "Starting simulation..."
        db.commit()

        _rollplan_jobs[rollplan_id] = {
            "status": "running",
            "progress": 0,
            "message": "Starting simulation...",
            "started_at": time.time(),
        }

        def progress_callback(pct: int, message: str):
            _rollplan_jobs[rollplan_id]["progress"] = pct
            _rollplan_jobs[rollplan_id]["message"] = message
            roll_plan.progress = pct
            roll_plan.progress_message = message
            db.commit()

        def cancel_check() -> bool:
            return _rollplan_jobs.get(rollplan_id, {}).get("status") == "cancelled"

        result = rollplan_service.run_simulation(
            db=db,
            roll_plan=roll_plan,
            progress_callback=progress_callback,
            cancel_check=cancel_check,
            ga_pop_size=ga_pop_size,
            ga_generations=ga_generations,
            confirm_shortfall=confirm_shortfall,
        )

        if result == "needs_confirmation":
            _rollplan_jobs[rollplan_id].update({
                "status": "needs_confirmation",
                "progress": 0,
                "message": roll_plan.roll_adjustment_message or "Roll shortfall detected",
            })
            db.commit()
        elif cancel_check():
            roll_plan.status = RollPlanStatus.cancelled
            roll_plan.progress_message = "Cancelled by user"
            _rollplan_jobs[rollplan_id].update({
                "status": "cancelled", "message": "Cancelled by user",
            })
        else:
            elapsed = time.time() - _rollplan_jobs[rollplan_id]["started_at"]
            msg = f"Complete in {elapsed:.0f}s"
            _rollplan_jobs[rollplan_id].update({
                "status": "completed", "progress": 100, "message": msg,
            })

        db.commit()

    except Exception as e:
        traceback.print_exc()
        elapsed = time.time() - _rollplan_jobs.get(rollplan_id, {}).get("started_at", time.time())
        err_msg = f"Failed after {elapsed:.0f}s: {str(e)}"

        _rollplan_jobs[rollplan_id] = {
            "status": "failed",
            "progress": _rollplan_jobs.get(rollplan_id, {}).get("progress", 0),
            "message": err_msg,
        }

        try:
            roll_plan = db.query(RollPlan).filter(RollPlan.id == rollplan_id).first()
            if roll_plan:
                roll_plan.status = RollPlanStatus.failed
                roll_plan.error_message = err_msg
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------


@router.post("", response_model=dict)
async def create_rollplan(
    request: RollPlanCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create a roll plan (config only). Upload rolls and start simulation separately."""
    # Verify cutplan exists and belongs to user
    cutplan = db.query(Cutplan).join(Order).filter(
        Cutplan.id == request.cutplan_id,
        Order.customer_id == current_user.customer_id,
    ).first()
    if not cutplan:
        raise HTTPException(status_code=404, detail="Cutplan not found")

    if cutplan.status == CutplanStatus.superseded:
        raise HTTPException(status_code=400, detail="Cannot create roll plan for a superseded cutplan")

    # Map mode string
    mode_map = {
        "monte_carlo": RollPlanMode.monte_carlo,
        "ga": RollPlanMode.ga,
        "both": RollPlanMode.both,
    }
    mode = mode_map.get(request.mode, RollPlanMode.both)

    roll_plan = RollPlan(
        cutplan_id=request.cutplan_id,
        name=request.name or f"Roll Plan - {cutplan.name or 'untitled'}",
        color_code=request.color_code,
        mode=mode,
        num_simulations=request.num_simulations,
        min_reuse_length_yards=request.min_reuse_length_yards,
        pseudo_roll_avg_yards=request.pseudo_roll_avg_yards,
        pseudo_roll_delta_yards=request.pseudo_roll_delta_yards,
        waste_threshold_pct=max(0.1, request.waste_threshold_pct),
        pseudo_buffer_pct=max(0.0, request.pseudo_buffer_pct),
        status=RollPlanStatus.pending,
        input_type=RollInputType.pseudo,
    )
    db.add(roll_plan)
    db.commit()
    db.refresh(roll_plan)

    return {"id": roll_plan.id, "status": "pending", "message": "Roll plan created"}


@router.post("/{rollplan_id}/upload-rolls", response_model=RollUploadResponse)
async def upload_rolls(
    rollplan_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload roll inventory Excel for a roll plan."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status not in (RollPlanStatus.pending, RollPlanStatus.completed, RollPlanStatus.failed):
        raise HTTPException(status_code=400, detail="Cannot upload rolls while simulation is running")

    file_bytes = await file.read()
    try:
        records = rollplan_service.upload_rolls(db, rollplan_id, file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    lengths = [r.length_yards for r in records]
    return RollUploadResponse(
        roll_plan_id=rollplan_id,
        rolls_count=len(records),
        total_length_yards=round(sum(lengths), 2),
        avg_length_yards=round(statistics.mean(lengths), 2) if lengths else 0,
        median_length_yards=round(statistics.median(lengths), 2) if lengths else 0,
        min_length_yards=round(min(lengths), 2) if lengths else 0,
        max_length_yards=round(max(lengths), 2) if lengths else 0,
        rolls=[FabricRollResponse.model_validate(r) for r in records],
    )


@router.post("/{rollplan_id}/simulate", response_model=dict)
async def start_simulation(
    rollplan_id: str,
    background_tasks: BackgroundTasks,
    ga_pop_size: int = 30,
    ga_generations: int = 50,
    confirm_shortfall: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Start MC/GA simulation in the background.

    If real rolls have a shortfall and confirm_shortfall=False, returns
    status="needs_confirmation" with the shortfall message. Re-call with
    confirm_shortfall=True to add generated rolls and proceed.
    """
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status == RollPlanStatus.running:
        raise HTTPException(status_code=409, detail="Simulation already running")

    # Reset for re-run
    roll_plan.status = RollPlanStatus.pending
    roll_plan.progress = 0
    roll_plan.progress_message = "Queued..."
    roll_plan.error_message = None
    db.commit()

    background_tasks.add_task(
        execute_rollplan_job,
        rollplan_id=rollplan_id,
        ga_pop_size=ga_pop_size,
        ga_generations=ga_generations,
        confirm_shortfall=confirm_shortfall,
    )

    return {"id": rollplan_id, "status": "queued", "message": "Simulation started"}


@router.get("/{rollplan_id}/status", response_model=RollPlanStatusResponse)
async def get_status(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Poll simulation progress."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    # Prefer in-memory job state (more current), fall back to DB
    job = _rollplan_jobs.get(rollplan_id)
    if job:
        return RollPlanStatusResponse(
            id=rollplan_id,
            status=job.get("status", "unknown"),
            progress=job.get("progress", 0),
            message=job.get("message", ""),
        )

    return RollPlanStatusResponse(
        id=rollplan_id,
        status=roll_plan.status.value if roll_plan.status else "pending",
        progress=roll_plan.progress or 0,
        message=roll_plan.progress_message or "",
    )


@router.get("/{rollplan_id}", response_model=RollPlanResponse)
async def get_rollplan(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get full roll plan results."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)
    data = rollplan_service.build_response_data(db, roll_plan)
    return RollPlanResponse(**data)


@router.get("", response_model=List[RollPlanListItem])
async def list_rollplans(
    cutplan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List roll plans, optionally filtered by cutplan."""
    query = (
        db.query(RollPlan)
        .join(Cutplan, RollPlan.cutplan_id == Cutplan.id)
        .join(Order, Cutplan.order_id == Order.id)
        .filter(Order.customer_id == current_user.customer_id)
    )
    if cutplan_id:
        query = query.filter(RollPlan.cutplan_id == cutplan_id)

    plans = query.order_by(RollPlan.created_at.desc()).all()
    return [RollPlanListItem.model_validate(p) for p in plans]


@router.get("/{rollplan_id}/dockets", response_model=List[CutDocketResponse])
async def get_dockets(
    rollplan_id: str,
    source: str = "mc",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get cutting dockets from a completed roll plan.

    Query param `source`: "mc" for Monte Carlo best run, "ga" for GA optimizer.
    """
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status != RollPlanStatus.completed:
        raise HTTPException(status_code=400, detail="Simulation not completed yet")

    if source == "ga":
        dockets = roll_plan.ga_dockets
    else:
        dockets = roll_plan.mc_best_run_dockets

    if not dockets:
        return []

    return [CutDocketResponse(**d) for d in dockets]


@router.get("/{rollplan_id}/dockets/{cut_number}", response_model=CutDocketResponse)
async def get_single_docket(
    rollplan_id: str,
    cut_number: int,
    source: str = "mc",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get a single cutting docket by cut number."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status != RollPlanStatus.completed:
        raise HTTPException(status_code=400, detail="Simulation not completed yet")

    if source == "ga":
        dockets = roll_plan.ga_dockets
    else:
        dockets = roll_plan.mc_best_run_dockets

    if not dockets:
        raise HTTPException(status_code=404, detail="No dockets available")

    for d in dockets:
        if d.get("cut_number") == cut_number:
            return CutDocketResponse(**d)

    raise HTTPException(status_code=404, detail=f"Cut number {cut_number} not found")


@router.get("/{rollplan_id}/rolls", response_model=List[FabricRollResponse])
async def get_rolls(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get all rolls (real + pseudo) for a roll plan."""
    _get_rollplan_or_404(db, rollplan_id, current_user)
    rolls = (
        db.query(FabricRoll)
        .filter(FabricRoll.roll_plan_id == rollplan_id)
        .order_by(FabricRoll.roll_number)
        .all()
    )
    return [FabricRollResponse.model_validate(r) for r in rolls]


@router.post("/{rollplan_id}/cancel", response_model=dict)
async def cancel_simulation(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Cancel a running simulation."""
    _get_rollplan_or_404(db, rollplan_id, current_user)

    job = _rollplan_jobs.get(rollplan_id)
    if not job or job.get("status") != "running":
        raise HTTPException(status_code=400, detail="No running simulation to cancel")

    _rollplan_jobs[rollplan_id]["status"] = "cancelled"
    _rollplan_jobs[rollplan_id]["message"] = "Cancelling..."
    return {"message": "Cancellation requested", "status": "cancelled"}


@router.delete("/{rollplan_id}", response_model=dict)
async def delete_rollplan(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a roll plan and its rolls."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status == RollPlanStatus.running:
        raise HTTPException(status_code=400, detail="Cannot delete while simulation is running")

    db.delete(roll_plan)
    db.commit()
    return {"message": "Roll plan deleted"}


# ---------------------------------------------------------------------------
# Roll Plan Excel Export
# ---------------------------------------------------------------------------


@router.get("/{rollplan_id}/export-excel")
async def export_rollplan_excel(
    rollplan_id: str,
    source: str = "ga",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Export roll plan as Excel workbook:
      Sheet 1: Lay Plan (cut-by-cut expansion)
      Sheet 2: Roll Plan Summary (waste + docket overview)
      Sheet 3+: Batched docket sheets (10 per tab)
    """
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    from ...models import Pattern
    from ...models.cutplan import CutplanMarker
    from ...services.excel_export_service import (
        write_lay_plan_sheet,
        write_roll_plan_summary_sheet, write_docket_batch_sheet,
    )

    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status != RollPlanStatus.completed:
        raise HTTPException(status_code=400, detail="Simulation not completed yet")

    # Pick dockets by source
    if source == "ga":
        dockets = roll_plan.ga_dockets
    else:
        dockets = roll_plan.mc_best_run_dockets

    if not dockets:
        raise HTTPException(status_code=400, detail="No dockets available for export")

    # Load cutplan with markers + layouts
    cutplan = (
        db.query(Cutplan)
        .options(joinedload(Cutplan.markers).joinedload(CutplanMarker.layout))
        .filter(Cutplan.id == roll_plan.cutplan_id)
        .first()
    )
    if not cutplan:
        raise HTTPException(status_code=404, detail="Parent cutplan not found")

    order = db.query(Order).filter(Order.id == cutplan.order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # --- Gather context ---
    canonical_sizes = None
    pattern = None
    if order.pattern_id:
        pattern = db.query(Pattern).filter(Pattern.id == order.pattern_id).first()
        if pattern and pattern.available_sizes:
            canonical_sizes = pattern.available_sizes

    from ...models import CostConfig as CostConfigModel
    cost_config = db.query(CostConfigModel).filter(
        CostConfigModel.customer_id == current_user.customer_id
    ).first()
    max_ply_height = cost_config.max_ply_height if cost_config else 100

    # Collect sizes from order
    demand_sizes_set = set()
    seen_colors = set()
    for line in order.order_lines:
        if line.color_code in seen_colors:
            continue
        seen_colors.add(line.color_code)
        for sq in line.size_quantities:
            if sq.quantity > 0:
                demand_sizes_set.add(sq.size_code)

    if canonical_sizes:
        sizes = [s for s in canonical_sizes if s in demand_sizes_set]
        for s in sorted(demand_sizes_set):
            if s not in sizes:
                sizes.append(s)
    else:
        sizes = sorted(demand_sizes_set)

    # Fabric code for this roll plan's color
    fabric_code = ""
    for line in order.order_lines:
        if line.color_code == roll_plan.color_code:
            fabric_code = line.fabric_code
            break

    plan_max_ply = max_ply_height
    if cutplan.solver_config and isinstance(cutplan.solver_config, dict):
        plan_max_ply = cutplan.solver_config.get("max_ply_height", max_ply_height)

    # Header info for all sheets
    header_info = {
        "order_number": order.order_number or "",
        "style_number": order.style_number or "",
        "fabric_code": fabric_code,
        "color_code": roll_plan.color_code or "",
        "cutplan_name": cutplan.name or "Cutplan",
    }

    # --- Map simulator cut numbers to lay-plan global cut numbers ---
    # The lay plan numbers cuts globally across all colors; the simulator
    # numbers them within a single color.  Walk the same marker/color/ply
    # expansion the lay plan uses and record the global cut numbers that
    # belong to this roll plan's color.
    rp_color = roll_plan.color_code or ""
    global_cuts_for_color: list[int] = []
    global_cut = 0
    for marker in cutplan.markers:
        plies_by_color = marker.plies_by_color or {}
        if not plies_by_color:
            plies_by_color = {"ALL": marker.total_plies or 0}
        for color, color_plies in plies_by_color.items():
            remaining = color_plies
            while remaining > 0:
                lay_plies = min(remaining, plan_max_ply)
                remaining -= lay_plies
                global_cut += 1
                if color == rp_color or (not rp_color and color == "ALL"):
                    global_cuts_for_color.append(global_cut)

    # Remap docket cut_numbers to global lay plan numbers
    for i, d in enumerate(dockets):
        if i < len(global_cuts_for_color):
            d["cut_number"] = global_cuts_for_color[i]

    # --- Build workbook ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Lay Plan (foundation — promoted to first sheet)
    write_lay_plan_sheet(wb, "Lay Plan", sizes, cutplan.markers, plan_max_ply,
                         header_info=header_info)

    # Sheet 2: Roll Plan Summary (waste stats + docket overview)
    waste_data = {}
    if source == "ga":
        waste_data = {
            "total_fabric_yards": roll_plan.total_fabric_required or 0,
            "unusable_yards": roll_plan.ga_unusable_yards or 0,
            "endbit_yards": roll_plan.ga_endbit_yards or 0,
            "returnable_yards": roll_plan.ga_returnable_yards or 0,
            "real_waste_yards": roll_plan.ga_real_waste_yards or 0,
        }
    else:
        waste_data = {
            "total_fabric_yards": roll_plan.total_fabric_required or 0,
            "unusable_yards": roll_plan.mc_unusable_avg or 0,
            "endbit_yards": roll_plan.mc_endbit_avg or 0,
            "returnable_yards": roll_plan.mc_returnable_avg or 0,
            "real_waste_yards": roll_plan.mc_real_waste_avg or 0,
        }

    write_roll_plan_summary_sheet(wb, dockets, waste_data, header_info=header_info)

    # Sheets 3+: Docket batches (10 per tab)
    BATCH_SIZE = 10
    for batch_start in range(0, len(dockets), BATCH_SIZE):
        batch = dockets[batch_start:batch_start + BATCH_SIZE]
        first = batch[0].get("cut_number", batch_start + 1)
        last = batch[-1].get("cut_number", batch_start + len(batch))
        sheet_name = f"Dockets {first}-{last}"
        write_docket_batch_sheet(wb, sheet_name, batch, header_info=header_info)

    # Serialize
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"order_{order.order_number}_rollplan.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Tune Cutplan (Approach 2: roll-aware ILP re-optimization)
# ---------------------------------------------------------------------------


def execute_tune_job(
    rollplan_id: str,
    avg_roll_length_yards: float,
    roll_penalty_weight: float = 2.0,
):
    """Execute cutplan tuning in the background."""
    db = SessionLocal()
    try:
        roll_plan = db.query(RollPlan).filter(RollPlan.id == rollplan_id).first()
        if not roll_plan:
            _tune_jobs[rollplan_id] = {
                "status": "failed", "progress": 0,
                "message": "Roll plan not found", "new_cutplan_id": None,
            }
            return

        _tune_jobs[rollplan_id] = {
            "status": "running", "progress": 0,
            "message": "Starting roll-optimized ILP...", "new_cutplan_id": None,
            "started_at": time.time(),
        }

        def progress_callback(pct: int, message: str):
            _tune_jobs[rollplan_id]["progress"] = pct
            _tune_jobs[rollplan_id]["message"] = message

        def cancel_check() -> bool:
            return _tune_jobs.get(rollplan_id, {}).get("status") == "cancelled"

        new_cutplan_id = rollplan_service.tune_cutplan(
            db=db,
            roll_plan=roll_plan,
            avg_roll_length_yards=avg_roll_length_yards,
            roll_penalty_weight=roll_penalty_weight,
            progress_callback=progress_callback,
            cancel_check=cancel_check,
        )

        elapsed = time.time() - _tune_jobs[rollplan_id]["started_at"]
        _tune_jobs[rollplan_id].update({
            "status": "completed", "progress": 100,
            "message": f"Tuning complete in {elapsed:.0f}s",
            "new_cutplan_id": new_cutplan_id,
        })

    except Exception as e:
        traceback.print_exc()
        _tune_jobs[rollplan_id] = {
            "status": "failed",
            "progress": _tune_jobs.get(rollplan_id, {}).get("progress", 0),
            "message": f"Tuning failed: {str(e)}",
            "new_cutplan_id": None,
        }
    finally:
        db.close()


@router.post("/{rollplan_id}/tune", response_model=dict)
async def tune_cutplan(
    rollplan_id: str,
    request: TuneCutplanRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Start roll-optimized ILP tuning. Requires MC simulation to be completed."""
    roll_plan = _get_rollplan_or_404(db, rollplan_id, current_user)

    if roll_plan.status != RollPlanStatus.completed:
        raise HTTPException(status_code=400, detail="Simulation must be completed before tuning")

    if roll_plan.mc_endbit_avg is None:
        raise HTTPException(status_code=400, detail="Monte Carlo results required for tuning")

    # Check if already tuning
    existing = _tune_jobs.get(rollplan_id)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Tuning already in progress")

    # Derive avg_roll_length from pseudo config if not provided
    avg_roll_length = request.avg_roll_length_yards
    if avg_roll_length is None:
        avg_roll_length = roll_plan.pseudo_roll_avg_yards or 100.0

    background_tasks.add_task(
        execute_tune_job,
        rollplan_id=rollplan_id,
        avg_roll_length_yards=avg_roll_length,
        roll_penalty_weight=request.roll_penalty_weight,
    )

    return {"status": "started", "message": "Roll-optimized ILP tuning started"}


@router.get("/{rollplan_id}/tune-status", response_model=TuneStatusResponse)
async def get_tune_status(
    rollplan_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Poll tuning progress."""
    _get_rollplan_or_404(db, rollplan_id, current_user)

    job = _tune_jobs.get(rollplan_id)
    if not job:
        return TuneStatusResponse(
            status="idle", progress=0, message="No tuning job found",
        )

    return TuneStatusResponse(
        status=job.get("status", "unknown"),
        progress=job.get("progress", 0),
        message=job.get("message", ""),
        new_cutplan_id=job.get("new_cutplan_id"),
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _get_rollplan_or_404(db: Session, rollplan_id: str, current_user: User) -> RollPlan:
    """Load roll plan with ownership check via cutplan→order→customer."""
    roll_plan = (
        db.query(RollPlan)
        .join(Cutplan, RollPlan.cutplan_id == Cutplan.id)
        .join(Order, Cutplan.order_id == Order.id)
        .filter(
            RollPlan.id == rollplan_id,
            Order.customer_id == current_user.customer_id,
        )
        .first()
    )
    if not roll_plan:
        raise HTTPException(status_code=404, detail="Roll plan not found")
    return roll_plan
