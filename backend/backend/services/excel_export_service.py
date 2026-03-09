"""
Excel export service — styles, helpers, charts, SVG→PNG, and lay plan generation
for the enhanced cutplan Excel report.
"""
import io
import math
import re
import xml.etree.ElementTree as ET
from typing import List, Dict, Tuple, Optional, Any

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import PieChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12, color="333399")

HEADER_FILL = PatternFill(start_color="E8E0D5", end_color="E8E0D5", fill_type="solid")
CUTPLAN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Efficiency thresholds
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Alternating row fill
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Size column header
SIZE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Cost section
COST_LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
COST_TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Totals row
TOTAL_ROW_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
TOTAL_ROW_FONT = Font(bold=True, size=11, color="FFFFFF")

# Marker detail header
MARKER_DETAIL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

CENTER = Alignment(horizontal="center")
DOLLAR_FMT = '$#,##0.00'
PCT_FMT = '0.0'

# PNG render resolution for marker images
MARKER_IMAGE_WIDTH = 900


# ---------------------------------------------------------------------------
# Helper: efficiency fill
# ---------------------------------------------------------------------------
def efficiency_fill(eff_pct: float) -> PatternFill:
    """Return green/amber/red fill based on efficiency percentage."""
    if eff_pct >= 80:
        return GREEN_FILL
    elif eff_pct >= 75:
        return AMBER_FILL
    return RED_FILL


# ---------------------------------------------------------------------------
# Helper: SVG → PNG bytes
# ---------------------------------------------------------------------------
def svg_to_png_bytes(svg_str: str, target_width: int = MARKER_IMAGE_WIDTH) -> Optional[bytes]:
    """
    Render a simple SVG (rect + polygon elements) to PNG bytes using PIL.
    Returns None on any failure (graceful degradation).
    """
    try:
        from PIL import Image, ImageDraw

        root = ET.fromstring(svg_str)
        ns = ""
        # Handle SVG namespace
        if root.tag.startswith("{"):
            ns = root.tag.split("}")[0] + "}"

        # Parse viewBox
        vb = root.get("viewBox", "")
        if not vb:
            w_attr = root.get("width", "600")
            h_attr = root.get("height", "200")
            vb_x, vb_y = 0.0, 0.0
            vb_w = float(w_attr.replace("px", "").replace("mm", ""))
            vb_h = float(h_attr.replace("px", "").replace("mm", ""))
        else:
            parts = vb.split()
            vb_x, vb_y, vb_w, vb_h = (float(p) for p in parts)

        if vb_w <= 0 or vb_h <= 0:
            return None

        scale = target_width / vb_w
        img_w = target_width
        img_h = max(1, int(vb_h * scale))

        img = Image.new("RGBA", (img_w, img_h), (255, 255, 255, 255))
        draw = ImageDraw.Draw(img)

        def transform(x: float, y: float) -> Tuple[float, float]:
            return ((x - vb_x) * scale, (y - vb_y) * scale)

        def parse_color(fill_str: str) -> Tuple[int, ...]:
            """Parse hex color or named color, return RGBA tuple."""
            if not fill_str or fill_str == "none":
                return (0, 0, 0, 0)
            fill_str = fill_str.strip()
            if fill_str.startswith("#"):
                h = fill_str.lstrip("#")
                if len(h) == 3:
                    h = "".join(c * 2 for c in h)
                if len(h) == 6:
                    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16), 255)
            # Common SVG colors
            colors = {
                "white": (255, 255, 255, 255),
                "black": (0, 0, 0, 255),
                "red": (255, 0, 0, 255),
                "green": (0, 128, 0, 255),
                "blue": (0, 0, 255, 255),
                "gray": (128, 128, 128, 255),
                "lightgray": (211, 211, 211, 255),
                "none": (0, 0, 0, 0),
            }
            return colors.get(fill_str.lower(), (200, 200, 200, 255))

        # Draw rectangles
        for rect in root.iter(f"{ns}rect"):
            x = float(rect.get("x", 0))
            y = float(rect.get("y", 0))
            w = float(rect.get("width", 0))
            h = float(rect.get("height", 0))
            fill = rect.get("fill", "#CCCCCC")
            stroke = rect.get("stroke", "none")
            tx1, ty1 = transform(x, y)
            tx2, ty2 = transform(x + w, y + h)
            color = parse_color(fill)
            if color[3] > 0:
                draw.rectangle([tx1, ty1, tx2, ty2], fill=color)
            if stroke and stroke != "none":
                draw.rectangle([tx1, ty1, tx2, ty2], outline=parse_color(stroke))

        # Draw polygons
        for poly in root.iter(f"{ns}polygon"):
            points_str = poly.get("points", "")
            if not points_str:
                continue
            coords = []
            for pair in points_str.strip().split():
                parts = pair.split(",")
                if len(parts) == 2:
                    coords.append(transform(float(parts[0]), float(parts[1])))
            if len(coords) >= 3:
                fill = poly.get("fill", "#AAAAAA")
                stroke = poly.get("stroke", "black")
                opacity = float(poly.get("opacity", poly.get("fill-opacity", "1")))
                color = parse_color(fill)
                if opacity < 1:
                    color = (color[0], color[1], color[2], int(opacity * 255))
                draw.polygon(coords, fill=color, outline=parse_color(stroke))

        # Draw paths (basic: only M/L/Z commands for simple polygons)
        for path_el in root.iter(f"{ns}path"):
            d = path_el.get("d", "")
            if not d:
                continue
            coords = []
            tokens = d.replace(",", " ").split()
            i = 0
            while i < len(tokens):
                cmd = tokens[i]
                if cmd in ("M", "L"):
                    i += 1
                    if i + 1 < len(tokens):
                        try:
                            coords.append(transform(float(tokens[i]), float(tokens[i + 1])))
                            i += 2
                        except ValueError:
                            i += 1
                elif cmd == "Z" or cmd == "z":
                    i += 1
                else:
                    try:
                        x_val = float(tokens[i])
                        if i + 1 < len(tokens):
                            y_val = float(tokens[i + 1])
                            coords.append(transform(x_val, y_val))
                            i += 2
                        else:
                            i += 1
                    except ValueError:
                        i += 1
            if len(coords) >= 3:
                fill = path_el.get("fill", "#AAAAAA")
                stroke = path_el.get("stroke", "black")
                color = parse_color(fill)
                draw.polygon(coords, fill=color, outline=parse_color(stroke))

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Write demand table
# ---------------------------------------------------------------------------
def write_demand_table(
    ws,
    row: int,
    sizes: List[str],
    fabric_color_demands: Dict[str, Dict[str, int]],
) -> int:
    """Write the demand table. Returns next available row."""
    ws.cell(row=row, column=1, value="DEMAND").font = HEADER_FONT
    row += 1

    # Header row
    ws.cell(row=row, column=1, value="Color").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    for i, size in enumerate(sizes):
        cell = ws.cell(row=row, column=2 + i, value=size)
        cell.font = HEADER_FONT
        cell.fill = SIZE_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    total_cell = ws.cell(row=row, column=2 + len(sizes), value="Total")
    total_cell.font = HEADER_FONT
    total_cell.fill = HEADER_FILL
    total_cell.border = THIN_BORDER
    row += 1

    # Color rows
    for color, demand in fabric_color_demands.items():
        ws.cell(row=row, column=1, value=color).border = THIN_BORDER
        for i, size in enumerate(sizes):
            qty = demand.get(size, 0)
            cell = ws.cell(row=row, column=2 + i, value=qty if qty else "")
            cell.border = THIN_BORDER
            cell.alignment = CENTER
        total = sum(demand.get(s, 0) for s in sizes)
        ws.cell(row=row, column=2 + len(sizes), value=total).border = THIN_BORDER
        row += 1

    return row + 2  # blank rows after demand


# ---------------------------------------------------------------------------
# Write marker table (no images — those go in marker details section)
# ---------------------------------------------------------------------------
def write_marker_table(
    ws,
    row: int,
    sizes: List[str],
    markers: list,
    perimeter_by_size: Dict[str, float],
    cost_params: Dict[str, float],
    max_ply_height: int = 100,
) -> Tuple[int, Dict[str, float]]:
    """
    Write the marker table with color-coded efficiency and zebra rows.
    Returns (next_row, cost_totals_dict).
    """
    DEFAULT_PERIMETER_CM = cost_params.get("default_perimeter_cm", 2540.0)
    fc_cost_per_yard = cost_params.get("fc_cost_per_yard", 3.0)
    spreading_cost_per_yard = cost_params.get("spreading_cost_per_yard", 0.00122)
    spreading_cost_per_ply = cost_params.get("spreading_cost_per_ply", 0.013)
    cutting_cost_per_cm = cost_params.get("cutting_cost_per_cm", 0.0)
    prep_cost_per_meter = cost_params.get("prep_cost_per_meter", 0.25)

    ws.cell(row=row, column=1, value="MARKERS").font = HEADER_FONT
    row += 1

    # Header
    marker_headers = ["Marker #"] + sizes + [
        "Bundles", "Efficiency %", "Length (yd)", "Plies", "Cuts", "Source",
    ]
    for i, h in enumerate(marker_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    row += 1

    # Accumulators
    totals_by_size = {s: 0 for s in sizes}
    total_plies_sum = 0
    total_cuts_sum = 0
    total_fabric_yards = 0.0
    total_cutting_cost = 0.0
    total_prep_cost = 0.0

    for m_idx, marker in enumerate(markers, 1):
        ratios = marker.ratio_str.split("-")
        bundles = sum(int(x) for x in ratios)

        has_cpu = marker.layout is not None
        source = "CPU" if has_cpu else "GPU"
        eff = (
            (marker.layout.utilization * 100)
            if has_cpu and marker.layout.utilization
            else ((marker.efficiency or 0) * 100)
        )
        length_yd = (
            marker.layout.length_yards
            if has_cpu and marker.layout.length_yards
            else (marker.length_yards or 0)
        )
        length_m = length_yd * 0.9144

        marker_plies = marker.total_plies or 0
        # Always recalculate cuts from current max_ply_height (DB value may be stale)
        marker_cuts = (marker_plies + max_ply_height - 1) // max_ply_height if marker_plies > 0 else 0

        # Zebra striping
        is_alt = m_idx % 2 == 0
        row_fill = ALT_ROW_FILL if is_alt else None

        def _style(cell, fill_override=None):
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if fill_override:
                cell.fill = fill_override
            elif row_fill:
                cell.fill = row_fill

        c = ws.cell(row=row, column=1, value=m_idx)
        _style(c)

        for i, size in enumerate(sizes):
            ratio_val = int(ratios[i]) if i < len(ratios) else 0
            cell = ws.cell(row=row, column=2 + i, value=ratio_val if ratio_val else "")
            _style(cell)
            totals_by_size[size] += ratio_val * marker_plies

        col_off = 2 + len(sizes)
        _style(ws.cell(row=row, column=col_off, value=bundles))

        # Efficiency cell with color coding
        eff_cell = ws.cell(row=row, column=col_off + 1, value=round(eff, 1))
        eff_cell.border = THIN_BORDER
        eff_cell.alignment = CENTER
        eff_cell.number_format = PCT_FMT
        eff_cell.fill = efficiency_fill(eff)

        _style(ws.cell(row=row, column=col_off + 2, value=round(length_yd, 2)))
        _style(ws.cell(row=row, column=col_off + 3, value=marker_plies))
        _style(ws.cell(row=row, column=col_off + 4, value=marker_cuts))
        _style(ws.cell(row=row, column=col_off + 5, value=source))

        total_plies_sum += marker_plies
        total_cuts_sum += marker_cuts
        total_fabric_yards += length_yd * marker_plies

        # Cutting cost
        marker_perim_cm = 0.0
        for i, size in enumerate(sizes):
            rv = int(ratios[i]) if i < len(ratios) else 0
            if rv > 0:
                marker_perim_cm += perimeter_by_size.get(size, DEFAULT_PERIMETER_CM) * rv
        total_cutting_cost += marker_perim_cm * marker_cuts * cutting_cost_per_cm

        # Prep cost
        total_prep_cost += length_m * marker_cuts * prep_cost_per_meter

        row += 1

    # Totals row (dark navy, white text)
    num_cols = len(marker_headers)
    for ci in range(1, 1 + num_cols):
        cell = ws.cell(row=row, column=ci)
        cell.fill = TOTAL_ROW_FILL
        cell.font = TOTAL_ROW_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_ROW_FONT
    ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
    for i, size in enumerate(sizes):
        cell = ws.cell(row=row, column=2 + i, value=totals_by_size[size])
        cell.font = TOTAL_ROW_FONT
        cell.fill = TOTAL_ROW_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    col_off = 2 + len(sizes)
    ws.cell(row=row, column=col_off + 3, value=total_plies_sum)
    ws.cell(row=row, column=col_off + 4, value=total_cuts_sum)
    row += 2

    # Compute costs
    spreading_cost = (total_fabric_yards * spreading_cost_per_yard) + (total_plies_sum * spreading_cost_per_ply)
    fabric_cost = total_fabric_yards * fc_cost_per_yard

    cost_totals = {
        "fabric_cost": fabric_cost,
        "spreading_cost": spreading_cost,
        "cutting_cost": total_cutting_cost,
        "prep_cost": total_prep_cost,
        "total_fabric_yards": total_fabric_yards,
        "total_plies": total_plies_sum,
        "total_cuts": total_cuts_sum,
    }
    return row, cost_totals


# ---------------------------------------------------------------------------
# Write cost breakdown
# ---------------------------------------------------------------------------
def write_cost_breakdown(
    ws,
    row: int,
    cost_totals: Dict[str, float],
    fabric_cost_per_yard: Optional[float] = None,
) -> Tuple[int, int]:
    """
    Write styled cost breakdown section with numeric values.
    Returns (next_row, cost_data_start_row) — cost_data_start_row is used by charts.
    """
    ws.cell(row=row, column=1, value="COST BREAKDOWN").font = HEADER_FONT
    if fabric_cost_per_yard is not None:
        info_cell = ws.cell(row=row, column=2, value=f"(Fabric: ${fabric_cost_per_yard:.2f}/yd)")
        info_cell.font = Font(italic=True, size=10, color="666666")
    row += 1

    cost_items = [
        ("Fabric Cost", cost_totals["fabric_cost"]),
        ("Spreading Cost", cost_totals["spreading_cost"]),
        ("Cutting Cost", cost_totals["cutting_cost"]),
        ("Prep Cost", cost_totals["prep_cost"]),
    ]
    total_cost = sum(v for _, v in cost_items)

    data_start = row

    for label, value in cost_items:
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.fill = COST_LABEL_FILL
        lbl_cell.border = THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.number_format = DOLLAR_FMT
        val_cell.border = THIN_BORDER
        row += 1

    # Total Cost row
    lbl_cell = ws.cell(row=row, column=1, value="Total Cost")
    lbl_cell.font = HEADER_FONT
    lbl_cell.fill = COST_TOTAL_FILL
    lbl_cell.border = THIN_BORDER
    val_cell = ws.cell(row=row, column=2, value=total_cost)
    val_cell.font = HEADER_FONT
    val_cell.number_format = DOLLAR_FMT
    val_cell.fill = COST_TOTAL_FILL
    val_cell.border = THIN_BORDER
    row += 1

    return row, data_start


# ---------------------------------------------------------------------------
# Write cost pie charts (two charts side by side)
# ---------------------------------------------------------------------------
def write_cost_charts(ws, data_start_row: int) -> None:
    """
    Add two pie charts beside cost data:
      1. "Fabric Cost vs Other Costs" — shows how dominant fabric is
      2. "Other Costs Breakdown" — zooms into spreading, cutting, prep

    We write helper data in columns E-F to feed the charts, since
    openpyxl PieChart reads from cell ranges.
    """
    # Read cost values from column B
    fabric_val = ws.cell(row=data_start_row, column=2).value or 0
    spreading_val = ws.cell(row=data_start_row + 1, column=2).value or 0
    cutting_val = ws.cell(row=data_start_row + 2, column=2).value or 0
    prep_val = ws.cell(row=data_start_row + 3, column=2).value or 0
    other_total = spreading_val + cutting_val + prep_val

    # --- Chart 1 data: columns E-F (Fabric vs Other) ---
    helper_col_lbl = 5   # E
    helper_col_val = 6   # F
    r = data_start_row
    ws.cell(row=r, column=helper_col_lbl, value="Fabric Cost")
    ws.cell(row=r, column=helper_col_val, value=fabric_val)
    ws.cell(row=r, column=helper_col_val).number_format = DOLLAR_FMT
    ws.cell(row=r + 1, column=helper_col_lbl, value="Other Costs")
    ws.cell(row=r + 1, column=helper_col_val, value=other_total)
    ws.cell(row=r + 1, column=helper_col_val).number_format = DOLLAR_FMT

    # Make helper cells small/light so they don't stand out
    for dr in range(2):
        ws.cell(row=r + dr, column=helper_col_lbl).font = Font(size=8, color="999999")
        ws.cell(row=r + dr, column=helper_col_val).font = Font(size=8, color="999999")

    chart1 = PieChart()
    chart1.title = "Fabric vs Other Costs"
    chart1.style = 10
    data1 = Reference(ws, min_col=helper_col_val, min_row=r, max_row=r + 1)
    labels1 = Reference(ws, min_col=helper_col_lbl, min_row=r, max_row=r + 1)
    chart1.add_data(data1, titles_from_data=False)
    chart1.set_categories(labels1)
    chart1.width = 13
    chart1.height = 10
    ws.add_chart(chart1, f"{get_column_letter(4)}{data_start_row}")

    # --- Chart 2 data: columns H-I (Other breakdown) ---
    helper_col_lbl2 = 8   # H
    helper_col_val2 = 9   # I
    ws.cell(row=r, column=helper_col_lbl2, value="Spreading")
    ws.cell(row=r, column=helper_col_val2, value=spreading_val)
    ws.cell(row=r, column=helper_col_val2).number_format = DOLLAR_FMT
    ws.cell(row=r + 1, column=helper_col_lbl2, value="Cutting")
    ws.cell(row=r + 1, column=helper_col_val2, value=cutting_val)
    ws.cell(row=r + 1, column=helper_col_val2).number_format = DOLLAR_FMT
    ws.cell(row=r + 2, column=helper_col_lbl2, value="Prep")
    ws.cell(row=r + 2, column=helper_col_val2, value=prep_val)
    ws.cell(row=r + 2, column=helper_col_val2).number_format = DOLLAR_FMT

    for dr in range(3):
        ws.cell(row=r + dr, column=helper_col_lbl2).font = Font(size=8, color="999999")
        ws.cell(row=r + dr, column=helper_col_val2).font = Font(size=8, color="999999")

    chart2 = PieChart()
    chart2.title = "Other Costs Breakdown"
    chart2.style = 10
    data2 = Reference(ws, min_col=helper_col_val2, min_row=r, max_row=r + 2)
    labels2 = Reference(ws, min_col=helper_col_lbl2, min_row=r, max_row=r + 2)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(labels2)
    chart2.width = 13
    chart2.height = 10

    # Place chart 2 to the right of chart 1
    anchor_col = get_column_letter(11)  # column K
    ws.add_chart(chart2, f"{anchor_col}{data_start_row}")


# ---------------------------------------------------------------------------
# Write marker details section (images after the full cutplan)
# ---------------------------------------------------------------------------
def write_marker_details(
    ws,
    row: int,
    sizes: List[str],
    markers: list,
    fabric_width_inches: Optional[float] = None,
) -> int:
    """
    Write a "MARKER DETAILS" section with per-marker header rows + PNG images.
    Header per marker:
        ID | Width | Sizes / Quantities | Bundles | Efficiency | Length | Time
    Then the PNG image below.
    Returns next available row.
    """
    ws.cell(row=row, column=1, value="MARKER DETAILS").font = TITLE_FONT
    row += 1

    # Column headers
    detail_headers = ["ID", "Width", "Sizes / Quantities", "Bundles", "Efficiency", "Length (yd)", "Time (s)"]
    for i, h in enumerate(detail_headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    row += 1

    width_str = f'{fabric_width_inches}"W' if fabric_width_inches else ""

    for m_idx, marker in enumerate(markers, 1):
        has_cpu = marker.layout is not None
        if not has_cpu or not marker.layout or not marker.layout.svg_preview:
            continue

        ratios = marker.ratio_str.split("-")
        bundles = sum(int(x) for x in ratios)
        eff = marker.layout.utilization * 100 if marker.layout.utilization else 0
        length_yd = marker.layout.length_yards or 0
        comp_time = marker.layout.computation_time_s or 0
        label = marker.marker_label or f"M{m_idx}"

        # Build "S:1, M:1, L:2" string
        size_qty_parts = []
        for i, size in enumerate(sizes):
            rv = int(ratios[i]) if i < len(ratios) else 0
            if rv > 0:
                size_qty_parts.append(f"{size}:{rv}")
        size_qty_str = ", ".join(size_qty_parts)

        # Data row with green header fill
        def _dc(cell):
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.fill = MARKER_DETAIL_FILL

        _dc(ws.cell(row=row, column=1, value=label))
        _dc(ws.cell(row=row, column=2, value=width_str))
        c = ws.cell(row=row, column=3, value=size_qty_str)
        c.border = THIN_BORDER
        c.fill = MARKER_DETAIL_FILL
        c.alignment = Alignment(horizontal="left")
        _dc(ws.cell(row=row, column=4, value=f"{bundles}bndl"))

        eff_cell = ws.cell(row=row, column=5, value=f"{eff:.1f}%")
        eff_cell.border = THIN_BORDER
        eff_cell.alignment = CENTER
        eff_cell.fill = efficiency_fill(eff)

        _dc(ws.cell(row=row, column=6, value=round(length_yd, 2)))
        _dc(ws.cell(row=row, column=7, value=round(comp_time, 1)))
        row += 1

        # Render and embed PNG
        png_bytes = svg_to_png_bytes(marker.layout.svg_preview, target_width=MARKER_IMAGE_WIDTH)
        if png_bytes:
            img_stream = io.BytesIO(png_bytes)
            img = XlImage(img_stream)
            # Display width in Excel points (roughly 7 columns wide)
            display_w = 650
            display_h = max(40, int(img.height * (display_w / max(img.width, 1))))
            img.width = display_w
            img.height = display_h
            anchor_col = get_column_letter(1)
            ws.add_image(img, f"{anchor_col}{row}")
            # Reserve rows (~18px per Excel row)
            rows_needed = max(2, math.ceil(display_h / 18))
            row += rows_needed
        else:
            row += 1  # blank row if image failed

    return row


# ---------------------------------------------------------------------------
# Write lay plan sheet
# ---------------------------------------------------------------------------
def write_lay_plan_sheet(
    wb,
    sheet_name: str,
    sizes: List[str],
    markers: list,
    max_ply_height: int = 100,
) -> None:
    """
    Create a new sheet expanding markers into per-lay rows
    based on max_ply_height.
    """
    # Excel forbids these characters in sheet names: \ / ? * [ ] :
    safe_name = re.sub(r'[\\/?*\[\]:]', '-', sheet_name)[:31]
    ws = wb.create_sheet(title=safe_name)

    # Header
    headers = ["Cut #", "Marker", "Color", "Ratio"] + sizes + ["Bundles", "Length (yd)", "Plies"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    row = 2
    cut_number = 0

    for marker in markers:
        ratios = marker.ratio_str.split("-")
        bundles = sum(int(x) for x in ratios)
        ratio_display = marker.ratio_str

        has_cpu = marker.layout is not None
        length_yd = (
            marker.layout.length_yards
            if has_cpu and marker.layout.length_yards
            else (marker.length_yards or 0)
        )

        label = marker.marker_label or f"M-{marker.ratio_str}"
        plies_by_color = marker.plies_by_color or {}

        if not plies_by_color:
            plies_by_color = {"ALL": marker.total_plies or 0}

        for color, color_plies in plies_by_color.items():
            remaining = color_plies
            while remaining > 0:
                lay_plies = min(remaining, max_ply_height)
                remaining -= lay_plies
                cut_number += 1

                is_alt = cut_number % 2 == 0
                row_fill = ALT_ROW_FILL if is_alt else None

                def _s(cell):
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                    if row_fill:
                        cell.fill = row_fill

                _s(ws.cell(row=row, column=1, value=cut_number))
                _s(ws.cell(row=row, column=2, value=label))
                _s(ws.cell(row=row, column=3, value=color))
                _s(ws.cell(row=row, column=4, value=ratio_display))

                for i, size in enumerate(sizes):
                    rv = int(ratios[i]) if i < len(ratios) else 0
                    _s(ws.cell(row=row, column=5 + i, value=rv if rv else ""))

                col_off = 5 + len(sizes)
                _s(ws.cell(row=row, column=col_off, value=bundles))
                c = ws.cell(row=row, column=col_off + 1, value=round(length_yd, 2))
                _s(c)
                _s(ws.cell(row=row, column=col_off + 2, value=lay_plies))

                row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 20)


# ---------------------------------------------------------------------------
# Write roll plan summary sheet
# ---------------------------------------------------------------------------
def write_roll_plan_summary_sheet(
    wb,
    dockets: List[Dict[str, Any]],
    waste_data: Dict[str, float],
) -> None:
    """
    Create a 'Roll Plan Summary' sheet with waste stats and docket overview.
    waste_data keys: total_fabric_yards, unusable_yards, endbit_yards,
                     returnable_yards, real_waste_yards
    """
    ws = wb.create_sheet(title="Roll Plan Summary")

    row = 1
    ws.cell(row=row, column=1, value="ROLL PLAN SUMMARY").font = TITLE_FONT
    row += 2

    # Waste stats block
    total_fabric = waste_data.get("total_fabric_yards", 0)
    unusable = waste_data.get("unusable_yards", 0)
    endbit = waste_data.get("endbit_yards", 0)
    returnable = waste_data.get("returnable_yards", 0)
    real_waste = waste_data.get("real_waste_yards", 0)
    waste_pct = (real_waste / total_fabric * 100) if total_fabric > 0 else 0

    ws.cell(row=row, column=1, value="Total Fabric Required:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=f"{total_fabric:.2f} yd")
    row += 1
    ws.cell(row=row, column=1, value="Total Waste:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=f"{real_waste:.2f} yd ({waste_pct:.1f}%)")
    row += 1
    ws.cell(row=row, column=1, value="Breakdown:").font = HEADER_FONT
    ws.cell(row=row, column=2, value=f"Unusable {unusable:.2f}  |  End-bit {endbit:.2f}  |  Returnable {returnable:.2f}")
    row += 2

    # Docket overview table
    headers = [
        "Cut #", "Marker", "Ratio", "Length (yd)", "Plies",
        "Planned", "Rolls", "Fabric (yd)", "End Bits (yd)",
    ]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    row += 1

    # Accumulators for totals
    total_plies = 0
    total_planned = 0
    total_rolls = 0
    total_fabric_used = 0.0
    total_end_bits = 0.0

    for d_idx, d in enumerate(dockets):
        plies = d.get("plies", 0)
        planned = d.get("plies_planned") if d.get("plies_planned") is not None else plies
        rolls_count = len(d.get("assigned_rolls", []))
        fabric_used = d.get("total_fabric_yards", 0)
        end_bits = d.get("total_end_bit_yards", 0)

        total_plies += plies
        total_planned += planned
        total_rolls += rolls_count
        total_fabric_used += fabric_used
        total_end_bits += end_bits

        is_alt = d_idx % 2 == 1
        is_shortfall = planned < plies

        def _s(cell, shortfall_highlight=False):
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if shortfall_highlight and is_shortfall:
                cell.fill = RED_FILL
            elif is_alt:
                cell.fill = ALT_ROW_FILL

        _s(ws.cell(row=row, column=1, value=d.get("cut_number", d_idx + 1)))
        _s(ws.cell(row=row, column=2, value=d.get("marker_label", "")))
        _s(ws.cell(row=row, column=3, value=d.get("ratio_str", "")))
        _s(ws.cell(row=row, column=4, value=round(d.get("marker_length_yards", 0), 2)))
        _s(ws.cell(row=row, column=5, value=plies))
        c = ws.cell(row=row, column=6, value=planned)
        _s(c, shortfall_highlight=True)
        _s(ws.cell(row=row, column=7, value=rolls_count))
        _s(ws.cell(row=row, column=8, value=round(fabric_used, 2)))
        _s(ws.cell(row=row, column=9, value=round(end_bits, 2)))
        row += 1

    # Totals row
    for ci in range(1, 1 + len(headers)):
        cell = ws.cell(row=row, column=ci)
        cell.fill = TOTAL_ROW_FILL
        cell.font = TOTAL_ROW_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_ROW_FONT
    ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
    ws.cell(row=row, column=5, value=total_plies)
    ws.cell(row=row, column=6, value=total_planned)
    ws.cell(row=row, column=7, value=total_rolls)
    ws.cell(row=row, column=8, value=round(total_fabric_used, 2))
    ws.cell(row=row, column=9, value=round(total_end_bits, 2))

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 22)


# ---------------------------------------------------------------------------
# Write individual docket sheet
# ---------------------------------------------------------------------------
def write_docket_sheet(
    wb,
    sheet_name: str,
    docket: Dict[str, Any],
) -> None:
    """
    Create a per-cut docket sheet with header block and roll assignment table.
    """
    safe_name = re.sub(r'[\\/?*\[\]:]', '-', sheet_name)[:31]
    ws = wb.create_sheet(title=safe_name)

    cut_num = docket.get("cut_number", "?")
    marker_label = docket.get("marker_label", "")
    marker_length = docket.get("marker_length_yards", 0)
    ratio_str = docket.get("ratio_str", "")
    plies = docket.get("plies", 0)
    planned = docket.get("plies_planned") if docket.get("plies_planned") is not None else plies

    # Header block
    row = 1
    ws.cell(row=row, column=1, value=f"Cut #{cut_num} — Marker: {marker_label}").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Marker Length: {marker_length:.2f} yd").font = HEADER_FONT
    ws.cell(row=row, column=3, value=f"Ratio: {ratio_str}").font = HEADER_FONT
    row += 1
    plies_str = f"Plies: {planned}/{plies}" if planned != plies else f"Plies: {plies}"
    ws.cell(row=row, column=1, value=plies_str).font = HEADER_FONT
    if planned < plies:
        ws.cell(row=row, column=1).fill = RED_FILL
    row += 2

    # Roll assignment table
    headers = ["Roll ID", "Roll Length (yd)", "Source", "Plies", "Fabric Used (yd)", "End Bit (yd)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    row += 1

    rolls = docket.get("assigned_rolls", [])
    total_plies = 0
    total_fabric = 0.0
    total_endbits = 0.0

    for r_idx, roll in enumerate(rolls):
        roll_id = roll.get("roll_id", "")
        roll_length = roll.get("roll_length_yards", 0)
        roll_plies = roll.get("plies_from_roll", 0)
        fabric_used = roll.get("fabric_used_yards", 0)
        end_bit = roll.get("end_bit_yards", 0)
        source = "Reuse" if "-bit" in str(roll_id) else "Fresh"

        total_plies += roll_plies
        total_fabric += fabric_used
        total_endbits += end_bit

        is_alt = r_idx % 2 == 1

        def _s(cell):
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if is_alt:
                cell.fill = ALT_ROW_FILL

        _s(ws.cell(row=row, column=1, value=roll_id))
        _s(ws.cell(row=row, column=2, value=round(roll_length, 2)))
        _s(ws.cell(row=row, column=3, value=source))
        _s(ws.cell(row=row, column=4, value=roll_plies))
        _s(ws.cell(row=row, column=5, value=round(fabric_used, 2)))
        _s(ws.cell(row=row, column=6, value=round(end_bit, 2)))
        row += 1

    # Totals row
    for ci in range(1, 1 + len(headers)):
        cell = ws.cell(row=row, column=ci)
        cell.fill = TOTAL_ROW_FILL
        cell.font = TOTAL_ROW_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_ROW_FONT
    ws.cell(row=row, column=1).fill = TOTAL_ROW_FILL
    ws.cell(row=row, column=4, value=total_plies)
    ws.cell(row=row, column=5, value=round(total_fabric, 2))
    ws.cell(row=row, column=6, value=round(total_endbits, 2))

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 22)
